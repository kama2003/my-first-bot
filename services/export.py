import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from models.calculation import Calculation
from utils.time_parser import minutes_to_str


def export_to_excel(calculations: list[Calculation]) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "История расчётов"

    headers = [
        "№", "Дата", "Материал", "Вес (г)", "Время печати",
        "Пластик ₽", "Электричество ₽", "Амортизация ₽", "Доп. расходы ₽",
        "Брак ₽", "Себестоимость ₽", "Цена продажи ₽"
    ]

    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row, calc in enumerate(calculations, 2):
        ws.cell(row=row, column=1, value=row - 1)
        ws.cell(row=row, column=2, value=calc.created_at.strftime("%d.%m.%Y %H:%M") if calc.created_at else "")
        ws.cell(row=row, column=3, value=calc.material)
        ws.cell(row=row, column=4, value=calc.weight)
        ws.cell(row=row, column=5, value=minutes_to_str(calc.print_time))
        ws.cell(row=row, column=6, value=calc.cost_filament)
        ws.cell(row=row, column=7, value=calc.cost_electricity)
        ws.cell(row=row, column=8, value=calc.cost_amortization)
        ws.cell(row=row, column=9, value=calc.cost_extra)
        ws.cell(row=row, column=10, value=calc.cost_defect)
        ws.cell(row=row, column=11, value=calc.total_cost)
        ws.cell(row=row, column=12, value=calc.sale_price)

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
